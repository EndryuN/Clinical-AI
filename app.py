# app.py
import os

# Load .env BEFORE any other imports so API keys are available
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _key, _val = _line.split('=', 1)
                os.environ[_key.strip()] = _val.strip()

import threading
from datetime import datetime
from flask import Flask, render_template, request, jsonify, Response, send_file, stream_with_context
from models import ExtractionSession, PatientBlock, FieldResult
from parser.docx_parser import parse_docx, get_raw_text
from extractor.llm_client import (check_ollama, generate, get_backend, set_backend,
    check_ollama_available, check_claude_available, list_ollama_models, get_ollama_model, set_ollama_model, SUGGESTED_MODELS)
from extractor.prompt_builder import build_prompt, build_all_prompts
from extractor.clinical_context import get_context_for_group
from extractor.response_parser import parse_llm_response
from extractor.regex_extractor import regex_extract, assign_unique_id
from extractor.coverage import compute_coverage
from extractor.preview_renderer import render_patient_preview
from export.excel_writer import write_excel
from config import get_groups, get_all_fields
from audit import log_event
import json

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'data')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global session (one at a time)
session = ExtractionSession()


@app.route('/')
def index():
    return render_template('index.html',
                           session_active=bool(session.patients),
                           current_backend=get_backend(),
                           ollama_available=check_ollama_available(),
                           claude_available=check_claude_available(),
                           ollama_models=list_ollama_models(),
                           suggested_models=SUGGESTED_MODELS,
                           current_ollama_model=get_ollama_model())


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    filename = file.filename.lower()

    if not (filename.endswith('.docx') or filename.endswith('.xlsx')):
        return jsonify({"error": "Only .docx and .xlsx files are supported"}), 400

    # Save file — clean up previous uploads first
    import time, glob
    for old in glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], '*_*.*')):
        try:
            os.remove(old)
        except OSError:
            pass
    safe_name = f"{int(time.time())}_{file.filename}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
    file.save(file_path)

    session.file_name = safe_name
    session.upload_time = datetime.now().isoformat()

    try:
        if filename.endswith('.xlsx'):
            # Import previously exported Excel — skip extraction
            patients = _import_excel(file_path)
            session.patients = patients
            session.status = 'complete'
            session.progress['total'] = len(patients)
            session.progress['current_patient'] = len(patients)

            log_event('import_excel', file_name=file.filename, patients_imported=len(patients))

            return jsonify({
                "status": "ok",
                "patients_detected": len(patients),
                "imported": True,
                "patient_list": [
                    {"id": p.id, "unique_id": p.unique_id, "initials": p.initials, "nhs_number": p.nhs_number}
                    for p in patients
                ]
            })
        else:
            # Parse .docx and detect patients
            session.status = 'parsing'
            patients = parse_docx(file_path)
            session.patients = patients
            session.status = 'parsed'
            session.progress['total'] = len(patients)

            # Render patient preview images (non-fatal if Pillow unavailable)
            # log_event is already used throughout app.py with signature (event_name, **kwargs)
            try:
                ts = safe_name.split('_')[0]
                preview_dir = os.path.join(app.static_folder, 'previews', ts)
                os.makedirs(preview_dir, exist_ok=True)
                for p in patients:
                    render_patient_preview(p, preview_dir)
            except Exception as preview_err:
                log_event('preview_render_error', error=str(preview_err))

            log_event('upload', file_name=file.filename, patients_detected=len(patients))

            return jsonify({
                "status": "ok",
                "patients_detected": len(patients),
                "imported": False,
                "patient_list": [
                    {"id": p.id, "unique_id": p.unique_id, "initials": p.initials, "nhs_number": p.nhs_number}
                    for p in patients
                ]
            })
    except Exception as e:
        session.status = 'idle'
        return jsonify({"error": str(e)}), 500


def _import_excel(file_path: str) -> list:
    """Import a previously exported Excel file back into PatientBlock objects.
    Supports both new format (with RawCells sheet + confidence_basis column)
    and legacy format (with old confidence column only).
    """
    import os, time
    from openpyxl import load_workbook

    all_fields = get_all_fields()
    groups = get_groups()

    wb = load_workbook(file_path)

    # ── Detect format ──
    has_rawcells = "RawCells" in wb.sheetnames
    is_new_format = False

    # ── Read Metadata sheet ──
    meta_lookup = {}   # (unique_id_or_pid, field_key) -> dict
    coverage_lookup = {}  # unique_id -> float (coverage_pct)

    if "Metadata" in wb.sheetnames:
        ws_meta = wb["Metadata"]
        # Row 1 = SOURCE_FILE
        if ws_meta.cell(row=1, column=1).value == "SOURCE_FILE":
            session.file_name = ws_meta.cell(row=1, column=2).value or ""
        # Row 2 = headers — read by name
        header_row = [ws_meta.cell(row=2, column=c).value for c in range(1, ws_meta.max_column + 1)]
        header_row = [h for h in header_row if h]

        def _col(name):
            try:
                return header_row.index(name)
            except ValueError:
                return None

        def _col_first(*names):
            for n in names:
                c = _col(n)
                if c is not None:
                    return c
            return None

        is_new_format = 'confidence_basis' in header_row

        pid_col    = _col_first('unique_id', 'patient_id')
        if pid_col is None:
            pid_col = 0
        fkey_col   = _col_first('field_key')
        if fkey_col is None:
            fkey_col = 1
        cbasis_col = _col('confidence_basis')
        conf_col   = _col('confidence')
        reason_col = _col('reason')
        scrow_col  = _col('source_cell_row')
        sccol_col  = _col('source_cell_col')
        snip_col   = _col('source_snippet')
        edited_col = _col('edited')
        orig_col   = _col('original_value')
        cpct_col   = _col('coverage_pct')
        # Legacy: source_cell as JSON
        sc_json_col = _col('source_cell')

        _LEGACY_MAP = {'high': 'structured_verbatim', 'medium': 'freeform_verbatim',
                       'low': 'freeform_inferred', 'none': 'absent'}

        for row in ws_meta.iter_rows(min_row=3, values_only=True):
            row = list(row)
            pid = str(row[pid_col]) if pid_col is not None and len(row) > pid_col and row[pid_col] else None
            fkey = str(row[fkey_col]) if fkey_col is not None and len(row) > fkey_col and row[fkey_col] else None
            if not pid or not fkey:
                continue

            if is_new_format and cbasis_col is not None:
                cb = row[cbasis_col] or 'absent'
            elif conf_col is not None:
                cb = _LEGACY_MAP.get(str(row[conf_col] or ''), 'absent')
            else:
                cb = 'structured_verbatim'

            source_cell = None
            if scrow_col is not None and sccol_col is not None:
                sc_r = row[scrow_col] if len(row) > scrow_col else None
                sc_c = row[sccol_col] if len(row) > sccol_col else None
                if sc_r is not None and sc_c is not None:
                    source_cell = {"row": int(sc_r), "col": int(sc_c)}
            elif sc_json_col is not None and len(row) > sc_json_col and row[sc_json_col]:
                try:
                    source_cell = json.loads(row[sc_json_col])
                except (json.JSONDecodeError, TypeError):
                    pass

            snippet = None
            if snip_col is not None and len(row) > snip_col and row[snip_col]:
                snippet = str(row[snip_col])

            meta_lookup[(pid, fkey)] = {
                "confidence_basis": cb,
                "reason": str(row[reason_col]) if reason_col is not None and len(row) > reason_col and row[reason_col] else '',
                "source_cell": source_cell,
                "source_snippet": snippet,
                "edited": str(row[edited_col]).lower() == 'true' if edited_col is not None and len(row) > edited_col else False,
                "original_value": str(row[orig_col]) if orig_col is not None and len(row) > orig_col and row[orig_col] else None,
            }
            if cpct_col is not None and len(row) > cpct_col and row[cpct_col] is not None and pid not in coverage_lookup:
                try:
                    coverage_lookup[pid] = float(row[cpct_col])
                except (ValueError, TypeError):
                    pass

    # ── Read RawCells sheet ──
    rawcells_lookup = {}
    coverage_map_lookup = {}

    if has_rawcells:
        ws_rc = wb["RawCells"]
        rc_headers = [ws_rc.cell(row=1, column=c).value for c in range(1, 6)]
        def _rccol(name):
            try: return rc_headers.index(name)
            except ValueError: return None
        uid_c = _rccol('unique_id') or 0
        row_c = _rccol('row') or 1
        col_c = _rccol('col') or 2
        txt_c = _rccol('text') or 3
        cjson_c = _rccol('coverage_json') or 4

        for row in ws_rc.iter_rows(min_row=2, values_only=True):
            row = list(row)
            pid = str(row[uid_c]) if row[uid_c] else None
            if not pid:
                continue
            rawcells_lookup.setdefault(pid, []).append({
                "row": int(row[row_c] or 0),
                "col": int(row[col_c] or 0),
                "text": str(row[txt_c] or ''),
            })
            if row[cjson_c]:
                try:
                    spans = json.loads(row[cjson_c])
                    cell_key = f"{int(row[row_c] or 0)},{int(row[col_c] or 0)}"
                    coverage_map_lookup.setdefault(pid, {})[cell_key] = spans
                except (json.JSONDecodeError, TypeError):
                    pass

    # ── Read Prototype V1 ──
    ws = wb.active
    has_uid_col = ws.cell(row=1, column=1).value == "unique_id"
    field_offset = 1 if has_uid_col else 0

    patients = []
    for row_idx in range(2, ws.max_row + 1):
        mrn_schema_col = next((f['excel_column'] for f in all_fields if f['key'] == 'mrn'), 3)
        nhs_schema_col = next((f['excel_column'] for f in all_fields if f['key'] == 'nhs_number'), 4)
        mrn_val = ws.cell(row=row_idx, column=mrn_schema_col + field_offset).value
        nhs_val = ws.cell(row=row_idx, column=nhs_schema_col + field_offset).value
        if not mrn_val and not nhs_val:
            continue

        unique_id = str(ws.cell(row=row_idx, column=1).value or '').strip() if has_uid_col else ''
        patient_id = str(mrn_val).strip() if mrn_val else f"patient_{row_idx - 1:03d}"
        lookup_key = unique_id or patient_id

        extractions = {}
        for group in groups:
            group_fields = {}
            for field in group['fields']:
                col = field['excel_column'] + field_offset
                cell_value = ws.cell(row=row_idx, column=col).value
                if cell_value is not None:
                    value = str(cell_value).strip()
                    meta = meta_lookup.get((lookup_key, field['key']),
                           meta_lookup.get((patient_id, field['key']), {}))
                    group_fields[field['key']] = FieldResult(
                        value=value,
                        confidence_basis=meta.get('confidence_basis', 'structured_verbatim'),
                        reason=meta.get('reason', ''),
                        source_cell=meta.get('source_cell'),
                        source_snippet=meta.get('source_snippet'),
                        edited=meta.get('edited', False),
                        original_value=meta.get('original_value'),
                    )
                else:
                    group_fields[field['key']] = FieldResult(value=None, confidence_basis='absent')
            extractions[group['name']] = group_fields

        demo = extractions.get("Demographics", {})
        initials = demo.get("initials", FieldResult()).value or ''
        nhs_number = demo.get("nhs_number", FieldResult()).value or ''
        mrn = demo.get("mrn", FieldResult()).value or patient_id

        raw_cells = rawcells_lookup.get(lookup_key, [])
        c_map = coverage_map_lookup.get(lookup_key, {})
        c_pct = coverage_lookup.get(lookup_key)

        patients.append(PatientBlock(
            id=mrn,
            unique_id=unique_id,
            initials=initials,
            nhs_number=nhs_number,
            raw_text="(imported from Excel)",
            extractions=extractions,
            raw_cells=raw_cells,
            coverage_map=c_map,
            coverage_pct=c_pct,
        ))

    wb.close()

    # ── Regenerate preview PNGs from raw_cells ──
    if has_rawcells and patients:
        try:
            ts = str(int(time.time()))
            preview_dir = os.path.join(app.static_folder, 'previews', ts)
            os.makedirs(preview_dir, exist_ok=True)
            for p in patients:
                if p.raw_cells:
                    render_patient_preview(p, preview_dir)
            session.file_name = f"{ts}_imported.xlsx"
        except Exception as preview_err:
            log_event('preview_render_error', error=str(preview_err))

    return patients


@app.route('/extract', methods=['POST'])
def extract():
    if session.status not in ('parsed', 'complete'):
        return jsonify({"error": "No document uploaded or already extracting"}), 400

    data = request.json or {}
    patient_limit = data.get('limit', None)  # Optional: limit number of patients to process
    concurrency = int(data.get('concurrency', 1))

    session.status = 'extracting'
    session.stop_requested = False
    session.concurrency = concurrency

    # Run extraction in background thread
    thread = threading.Thread(target=_run_extraction, args=(patient_limit, concurrency))
    thread.daemon = True
    thread.start()

    return jsonify({"status": "started"})


@app.route('/stop', methods=['POST'])
def stop_extraction():
    """Request to stop the background extraction thread."""
    session.stop_requested = True
    return jsonify({"status": "stop_requested"})


def _run_extraction(patient_limit=None, concurrency=1):
    import time
    from concurrent.futures import ThreadPoolExecutor

    groups = get_groups()
    llm_groups = [g for g in groups if g.get('llm_required', False)]
    patients_to_process = session.patients[:patient_limit] if patient_limit else session.patients

    if not patients_to_process:
        session.status = 'complete'
        session.progress['phase'] = 'complete'
        return

    session.progress['total'] = len(patients_to_process)
    session.progress['phase'] = 'llm'
    session.progress['regex_complete'] = 0
    session.progress['llm_complete'] = 0
    session.progress['llm_queue_size'] = len(patients_to_process)
    session.progress['current_patient'] = 0
    session.progress['patient_times'] = []
    session.progress['average_seconds'] = 0
    session.progress['active_patients'] = {}
    session.progress['completed_patients'] = []
    session.progress['start_time'] = time.time()

    _counter_lock = threading.Lock()
    llm_semaphore = threading.Semaphore(concurrency)

    def process_patient(patient):
        if session.stop_requested:
            return

        # --- Regex (inline, fast) ---
        for group in groups:
            results = regex_extract(patient.raw_text, group['name'], group['fields'], patient.raw_cells)
            if group.get('llm_required', False):
                for key, fr in results.items():
                    if fr.value is None:
                        results[key] = FieldResult(value=None, confidence_basis='absent')
            patient.extractions[group['name']] = results
        with _counter_lock:
            session.progress['regex_complete'] += 1

        if session.stop_requested:
            return

        # --- LLM (serial via semaphore, groups run sequentially per patient) ---
        patient_llm_groups = [
            g for g in llm_groups
            if any(fr.value is None for fr in patient.extractions.get(g['name'], {}).values())
        ]

        if not patient_llm_groups:
            conf = _confidence_summary(patient)
            with _counter_lock:
                session.progress['llm_complete'] += 1
                session.progress['current_patient'] = session.progress['llm_complete']
                session.progress['completed_patients'].append({
                    "id": patient.id, "initials": patient.initials,
                    "confidence_summary": conf, "seconds": 0, "llm_seconds": 0,
                })
            return

        queued_time = time.time()
        llm_processing_time = 0.0  # actual LLM time (excludes queue wait)
        session.progress['active_patients'][patient.id] = {
            "initials": patient.initials,
            "group": patient_llm_groups[0]['name'],
            "start": queued_time,
            "llm_start": None,  # set when first semaphore acquired
            "status": "queued",
            "has_context": bool(get_context_for_group(patient_llm_groups[0]['name'])),
            "groups_done": 0,
            "groups_total": len(patient_llm_groups),
        }

        for group in patient_llm_groups:
            if session.stop_requested:
                break
            session.progress['active_patients'][patient.id].update({
                'group': group['name'],
                'has_context': bool(get_context_for_group(group['name'])),
                'status': 'queued',
            })
            with llm_semaphore:
                if session.stop_requested:
                    break
                group_start = time.time()
                ap = session.progress['active_patients'][patient.id]
                ap['status'] = 'running'
                if ap['llm_start'] is None:
                    ap['llm_start'] = group_start
                try:
                    system_prompt, user_prompt = build_prompt(patient.raw_text, group)
                    raw_response = generate(user_prompt, system_prompt)
                    llm_results = parse_llm_response(raw_response, group, raw_cells=patient.raw_cells)
                    for key, llm_fr in llm_results.items():
                        current = patient.extractions[group['name']].get(key)
                        if current and current.value is None and llm_fr.value is not None:
                            _resolve_source_cell(patient, llm_fr)
                            llm_fr.reason = f"[LLM] {llm_fr.reason}"
                            patient.extractions[group['name']][key] = llm_fr
                except Exception as e:
                    log_event('llm_extraction_error', patient_id=patient.id, group=group['name'], error=str(e))
                llm_processing_time += time.time() - group_start
            session.progress['active_patients'][patient.id]['groups_done'] += 1

        session.progress['active_patients'].pop(patient.id, None)

        llm_seconds = round(llm_processing_time, 1)
        conf = _confidence_summary(patient)
        with _counter_lock:
            session.progress['llm_complete'] += 1
            session.progress['current_patient'] = session.progress['llm_complete']
            session.progress['patient_times'].append(llm_seconds)
            session.progress['average_seconds'] = round(
                sum(session.progress['patient_times']) / len(session.progress['patient_times']), 1
            )
            # Throughput: wall time elapsed / patients done
            wall_elapsed = time.time() - session.progress['start_time']
            completed_count = session.progress['llm_complete']
            session.progress['throughput_seconds'] = round(wall_elapsed / completed_count, 1) if completed_count else 0
            session.progress['completed_patients'].append({
                "id": patient.id, "initials": patient.initials,
                "confidence_summary": conf, "seconds": llm_seconds, "llm_seconds": llm_seconds,
            })

    with ThreadPoolExecutor(max_workers=max(1, min(concurrency, 3))) as ex:
        list(ex.map(process_patient, patients_to_process))

    # Assign unique_ids and compute coverage for all patients
    for i, patient in enumerate(session.patients):
        if not patient.unique_id:
            assign_unique_id(patient, patient.extractions.get("Demographics", {}), row_index=i)
        compute_coverage(patient)
    _deduplicate_unique_ids(session.patients)

    session.status = 'complete' if not session.stop_requested else 'stopped'
    session.progress['phase'] = 'complete'


@app.route('/progress')
def progress():
    def event_stream():
        import time
        while session.status == 'extracting' or session.status == 'complete' or session.status == 'stopped':
            # Send update if patient changed, group changed, or every 2 seconds for timer
            event_data = {
                "current_patient": session.progress['current_patient'],
                "total": session.progress['total'],
                "phase": session.progress.get('phase', 'idle'),
                "regex_complete": session.progress.get('regex_complete', 0),
                "llm_complete": session.progress.get('llm_complete', 0),
                "llm_queue_size": session.progress.get('llm_queue_size', 0),
                "active_patients": session.progress.get('active_patients', {}),
                "completed_patients": session.progress.get('completed_patients', []),
                "average_seconds": round(session.progress.get('average_seconds', 0), 1),
                "throughput_seconds": round(session.progress.get('throughput_seconds', 0), 1),
                "start_time": session.progress.get('start_time', 0),
                "status": session.status
            }
            yield f"data: {json.dumps(event_data)}\n\n"
            
            if session.status in ('complete', 'stopped'):
                break
                
            time.sleep(2)

    return Response(stream_with_context(event_stream()), mimetype='text/event-stream')


@app.route('/process')
def process_page():
    return render_template('process.html', session_active=bool(session.patients))


@app.route('/patients')
def get_patients():
    cancer_type = request.args.get('cancer_type', '')
    search = request.args.get('search', '').lower()

    # During extraction only expose patients whose LLM is done
    if session.status == 'extracting':
        completed_ids = {p['id'] for p in session.progress.get('completed_patients', [])}
        source = [p for p in session.patients if p.id in completed_ids]
    else:
        source = session.patients

    result = []
    for p in source:
        ct = _get_cancer_type(p)

        if cancer_type and ct != cancer_type:
            continue
        if search and search not in p.initials.lower() and search not in p.nhs_number:
            continue

        conf = _confidence_summary(p)
        result.append({
            "id": p.id,
            "unique_id": p.unique_id,
            "initials": p.initials,
            "nhs_number": p.nhs_number,
            "gender": _get_field_value(p, "Demographics", "gender"),
            "cancer_type": ct,
            "confidence_summary": conf,
            "coverage_pct": p.coverage_pct,
        })

    return jsonify({"patients": result})


@app.route('/patients/<patient_id>')
def get_patient(patient_id):
    patient = _find_patient(patient_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404

    extractions = {}
    for group_name, fields in patient.extractions.items():
        extractions[group_name] = {
            key: {
                "value": fr.value,
                "confidence": fr.confidence,
                "confidence_basis": fr.confidence_basis,
                "reason": fr.reason,
                "edited": fr.edited,
                "source_cell": fr.source_cell,
                "source_snippet": fr.source_snippet,
            }
            for key, fr in fields.items()
        }

    return jsonify({
        "id": patient.id,
        "unique_id": patient.unique_id,
        "initials": patient.initials,
        "nhs_number": patient.nhs_number,
        "raw_text": patient.raw_text,
        "raw_cells": patient.raw_cells,
        "extractions": extractions,
        "coverage_map": patient.coverage_map,
        "coverage_pct": patient.coverage_pct,
        "coverage_stats": patient.coverage_stats,
    })


@app.route('/patient/<patient_id>/preview')
def patient_preview(patient_id):
    """Return rendered image URL and cell coordinate map for the patient."""
    patient = next(
        (p for p in session.patients
         if p.unique_id == patient_id or p.id == patient_id),
        None
    )
    if not patient:
        return jsonify({"error": "not found"}), 404
    if not session.file_name:
        return jsonify({"error": "no file"}), 404
    ts = session.file_name.split('_')[0]
    file_id = patient.unique_id if patient.unique_id else patient.id
    json_path = os.path.join(app.static_folder, 'previews', ts, f'{file_id}.json')
    if not os.path.exists(json_path):
        return jsonify({"error": "preview not available"}), 404
    with open(json_path) as f:
        coords = json.load(f)
    return jsonify({
        "image_url": f"/static/previews/{ts}/{file_id}.png",
        "coords": coords,
        "coverage_map": patient.coverage_map,
        "coverage_pct": patient.coverage_pct,
        "coverage_stats": patient.coverage_stats,
    })


@app.route('/patients/<patient_id>/fields', methods=['PUT'])
def edit_field(patient_id):
    patient = _find_patient(patient_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404

    data = request.json
    group = data.get('group')
    field_key = data.get('field')
    new_value = data.get('value')

    if group not in patient.extractions or field_key not in patient.extractions[group]:
        return jsonify({"error": "Field not found"}), 404

    fr = patient.extractions[group][field_key]
    old_value = fr.value

    if not fr.edited:
        fr.original_value = old_value
    fr.value = new_value
    fr.edited = True
    fr.confidence_basis = "edited"

    log_event('manual_edit',
              patient_id=patient.nhs_number,
              group=group, field=field_key,
              old_value=old_value, new_value=new_value)

    return jsonify({"status": "ok", "old_value": old_value, "new_value": new_value})


@app.route('/patients/<patient_id>/re-extract', methods=['POST'])
def re_extract(patient_id):
    patient = _find_patient(patient_id)
    if not patient:
        return jsonify({"error": "Patient not found"}), 404

    data = request.json or {}
    target_groups = data.get('groups', [g['name'] for g in get_groups()])

    def _do_re_extract():
        groups = get_groups()
        for group in groups:
            if group['name'] in target_groups:
                try:
                    # Phase 1: regex
                    results = regex_extract(patient.raw_text, group['name'], group['fields'], patient.raw_cells)
                    # Phase 2: LLM for gaps in LLM groups
                    if group.get('llm_required', False):
                        gaps = sum(1 for fr in results.values() if fr.value is None)
                        if gaps > 0:
                            system_prompt, user_prompt = build_prompt(patient.raw_text, group)
                            raw_response = generate(user_prompt, system_prompt)
                            llm_results = parse_llm_response(raw_response, group, raw_cells=patient.raw_cells)
                            for key, llm_fr in llm_results.items():
                                if key in results and results[key].value is None and llm_fr.value is not None:
                                    _resolve_source_cell(patient, llm_fr)
                                    llm_fr.reason = f"[LLM] {llm_fr.reason}"
                                    results[key] = llm_fr
                    patient.extractions[group['name']] = results
                except Exception as e:
                    log_event('re_extract_error', group=group['name'], error=str(e))
                    pass  # LLM call failed — regex results preserved for non-llm_required groups

    thread = threading.Thread(target=_do_re_extract)
    thread.daemon = True
    thread.start()

    return jsonify({"status": "started"})


@app.route('/status')
def status():
    return jsonify({
        "status": session.status,
        "current": session.progress['current_patient'],
        "total": session.progress['total'],
        "phase": session.progress.get('phase', 'idle'),
    })


@app.route('/export')
def export():
    if session.status not in ('complete', 'stopped') or not session.patients:
        return jsonify({"error": "No data to export"}), 400

    output_path = os.path.join(app.config['UPLOAD_FOLDER'], 'export.xlsx')
    write_excel(session.patients, output_path, source_name=session.file_name)

    log_event('export', patients_exported=len(session.patients), format='xlsx')

    return send_file(output_path,
                     download_name='mdt_extraction.xlsx',
                     as_attachment=True)


@app.route('/analytics')
def analytics_data():
    if not session.patients:
        return jsonify({})

    cancer_types = {}
    treatments = {}
    confidence = {"high": 0, "medium": 0, "low": 0}

    for p in session.patients:
        ct = _get_cancer_type(p)
        cancer_types[ct] = cancer_types.get(ct, 0) + 1

        treat = _get_field_value(p, "MDT", "first_mdt_treatment")
        if treat:
            # Extract key treatment keywords from the free-text outcome
            for keyword in _extract_treatment_keywords(treat):
                treatments[keyword] = treatments.get(keyword, 0) + 1

        for fields in p.extractions.values():
            for fr in fields.values():
                if fr.confidence in confidence:  # skip "none" (null/absent)
                    confidence[fr.confidence] += 1

    return jsonify({
        "cancer_types": cancer_types,
        "treatments": treatments,
        "confidence": confidence
    })


@app.route('/analytics/column/<field_key>')
def column_stats(field_key):
    """Compute statistics for a specific column across all patients."""
    import re
    from datetime import date

    is_dob = (field_key == 'dob')

    values = []
    numeric_values = []
    value_counts = {}

    for p in session.patients:
        for group_name, fields in p.extractions.items():
            if field_key in fields:
                v = fields[field_key].value
                if v is not None and v.lower() not in ('missing', 'n/a', ''):
                    if is_dob:
                        age = _dob_to_age(v)
                        if age is not None:
                            display = str(age)
                            values.append(display)
                            numeric_values.append(float(age))
                            value_counts[display] = value_counts.get(display, 0) + 1
                    else:
                        values.append(v)
                        value_counts[v] = value_counts.get(v, 0) + 1
                        try:
                            numeric_values.append(float(re.sub(r'[^\d.\-]', '', v)))
                        except (ValueError, TypeError):
                            pass

    result = {
        "field": "Age (from DOB)" if is_dob else field_key,
        "total_patients": len(session.patients),
        "populated": len(values),
        "empty": len(session.patients) - len(values),
        "unique_values": len(set(values)),
        "value_distribution": dict(sorted(value_counts.items(), key=lambda x: -x[1])[:20]),
    }

    if numeric_values:
        n = len(numeric_values)
        mean = sum(numeric_values) / n
        variance = sum((x - mean) ** 2 for x in numeric_values) / n if n > 1 else 0
        std_dev = variance ** 0.5
        result["numeric"] = True
        result["mean"] = round(mean, 2)
        result["std_dev"] = round(std_dev, 2)
        result["min"] = round(min(numeric_values), 2)
        result["max"] = round(max(numeric_values), 2)
        result["count"] = n
    else:
        result["numeric"] = False

    return jsonify(result)


def _dob_to_age(dob_str: str):
    """Convert a DOB string to age in years (for analytics only)."""
    from datetime import date
    import re
    try:
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', dob_str)
        if m:
            born = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        else:
            m = re.match(r'(\d{4})-(\d{2})-(\d{2})', dob_str)
            if m:
                born = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            else:
                return None
        today = date.today()
        return today.year - born.year - ((today.month, today.day) < (born.month, born.day))
    except (ValueError, TypeError):
        return None


@app.route('/analytics-page')
def analytics_page():
    return render_template('analytics.html', session_active=bool(session.patients))


@app.route('/schema')
def schema():
    """Return schema groups with colours and field keys for the frontend."""
    groups = get_groups()
    return jsonify([{
        "name": g['name'],
        "color": g.get('color', '#D9D9D9'),
        "fields": [f['key'] for f in g['fields']]
    } for g in groups])


@app.route('/review')
def review_page():
    return render_template('review.html', session_active=bool(session.patients))


@app.route('/audit')
def audit_trail():
    from audit import read_log
    return jsonify(read_log())


@app.route('/debug/raw-text')
def debug_raw_text():
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], session.file_name) if session.file_name else None
    if not file_path or not os.path.exists(file_path):
        return "No file uploaded", 404
    return f"<pre>{get_raw_text(file_path)}</pre>"


@app.route('/backend', methods=['GET', 'POST'])
def backend():
    """Get or set the LLM backend and model."""
    if request.method == 'POST':
        data = request.json or {}
        choice = data.get('backend', 'ollama')
        set_backend(choice)
        if 'ollama_model' in data:
            set_ollama_model(data['ollama_model'])
        return jsonify({"status": "ok", "backend": get_backend(), "ollama_model": get_ollama_model()})
    return jsonify({
        "backend": get_backend(),
        "ollama_available": check_ollama_available(),
        "claude_available": check_claude_available(),
        "ollama_model": get_ollama_model(),
        "ollama_models": list_ollama_models()
    })


@app.route('/link-source', methods=['POST'])
def link_source():
    """Upload a .docx to link previews/cells to an existing session (e.g. after Excel import)."""
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    if not session.patients:
        return jsonify({"error": "No active session to link to"}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.docx'):
        return jsonify({"error": "Only .docx files can be linked as source"}), 400

    # Save with timestamp
    import time
    safe_name = f"{int(time.time())}_{file.filename}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
    file.save(file_path)

    try:
        new_patients = parse_docx(file_path)
        linked_count = 0

        # Match by ID (Hospital Number) or NHS Number or Initials
        for p_existing in session.patients:
            match = None
            # 1. Try ID (Hospital Number)
            match = next((np for np in new_patients if np.id == p_existing.id), None)
            
            # 2. Try NHS Number
            if not match and p_existing.nhs_number:
                match = next((np for np in new_patients if np.nhs_number == p_existing.nhs_number), None)
            
            # 3. Try Initials (less reliable, but better than nothing if 1 & 2 fail)
            if not match and p_existing.initials:
                # Only match by initials if there's exactly one candidate in the new file
                candidates = [np for np in new_patients if np.initials == p_existing.initials]
                if len(candidates) == 1:
                    match = candidates[0]

            if match:
                p_existing.raw_cells = match.raw_cells
                p_existing.raw_text = match.raw_text
                linked_count += 1

        # Render previews for all patients (even if some didn't match, we try)
        ts = safe_name.split('_')[0]
        preview_dir = os.path.join(app.static_folder, 'previews', ts)
        os.makedirs(preview_dir, exist_ok=True)
        for p in session.patients:
            if p.raw_cells: # Only render if we have cells
                render_patient_preview(p, preview_dir)

        session.file_name = safe_name
        log_event('link_source', file_name=file.filename, matched=linked_count)

        return jsonify({
            "status": "ok",
            "matched": linked_count,
            "total": len(session.patients)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Helper functions
def _resolve_source_cell(patient, fr):
    """Search freeform cells (rows 4-7) for a cell containing fr.value and populate
    fr.source_cell/source_snippet.

    Note: uses substring matching on the normalised value — this is a best-effort
    approximation for LLM-extracted fields. For regex-extracted fields, source_cell
    is already set precisely by regex_extractor.py.
    """
    if not fr.value or not patient.raw_cells:
        return
    freeform = [c for c in patient.raw_cells if c.get('row', 0) in {4, 5, 6, 7}]
    for cell in freeform:
        if fr.value in cell["text"]:
            fr.source_cell = {"row": cell["row"], "col": cell["col"]}
            if fr.source_snippet is None:
                fr.source_snippet = fr.value[:200]
            return


def _find_patient(patient_id: str):
    for p in session.patients:
        if p.unique_id == patient_id or p.id == patient_id:
            return p
    return None


def _deduplicate_unique_ids(patients: list) -> None:
    """Ensure unique_id is unique within the batch. Append _b, _c... for collisions."""
    seen = {}
    suffix_chars = 'bcdefghijklmnopqrstuvwxyz'
    for patient in patients:
        uid = patient.unique_id
        if uid in seen:
            for ch in suffix_chars:
                candidate = f"{uid}_{ch}"
                if candidate not in seen:
                    patient.unique_id = candidate
                    seen[candidate] = True
                    break
        else:
            seen[uid] = True


def _get_field_value(patient, group_name, field_key):
    if group_name in patient.extractions and field_key in patient.extractions[group_name]:
        return patient.extractions[group_name][field_key].value
    return None


def _get_cancer_type(patient):
    import re
    # 1. Extract from Diagnosis line (e.g., "Diagnosis: ADENOCARCINOMA, NOT OTHERWISE SPECIFIED")
    m = re.search(r'Diagnosis:\s*([A-Za-z][A-Za-z\s\-]+?)(?:\s*[,()\n]|$)', patient.raw_text)
    if m:
        diag = m.group(1).strip()
        if not diag.upper().startswith('ICD'):
            diag = re.split(r',\s*', diag)[0].strip()
            return diag.title()
    # 2. Fallback: try the LLM-extracted biopsy_result
    biopsy = _get_field_value(patient, "Histology", "biopsy_result")
    if biopsy and biopsy.lower() not in ('missing', 'n/a', 'null'):
        return biopsy.split(',')[0].strip().title()
    # 3. No diagnosis yet
    return "Pending Diagnosis"


def _extract_treatment_keywords(text: str) -> list[str]:
    """Extract recognisable treatment categories from free-text MDT outcome."""
    import re
    text_lower = text.lower()
    found = []

    keywords = [
        ('TNT', r'\btnt\b'),
        ('Chemotherapy', r'chemo(?:therapy)?'),
        ('Radiotherapy', r'radio(?:therapy)?|chemoradio'),
        ('Short-course RT', r'short.?course'),
        ('Long-course CRT', r'long.?course'),
        ('Surgery', r'surgery|resection|hemicolectomy|colectomy|anterior resection|apr\b'),
        ('Watch & Wait', r'watch\s*(?:and|&)\s*wait'),
        ('Palliative', r'palliat'),
        ('MRI', r'\bmri\b'),
        ('CT scan', r'\bct\b(?!\.)|pet.?ct'),
        ('Papillon', r'papillon'),
        ('Immunotherapy', r'immuno(?:therapy)?|pembrolizumab|nivolumab'),
        ('Stoma', r'stoma|defunction'),
        ('Biopsy', r'biopsy'),
        ('Referred', r'refer|rediscuss|relist'),
    ]

    for label, pattern in keywords:
        if re.search(pattern, text_lower):
            found.append(label)

    return found if found else ['Other']


def _confidence_summary(patient):
    summary = {"high": 0, "medium": 0, "low": 0}
    for fields in patient.extractions.values():
        for fr in fields.values():
            if fr.confidence in summary:  # skip "none" (null/absent fields)
                summary[fr.confidence] += 1
    return summary


if __name__ == '__main__':
    import sys
    port = 5000
    if '--port' in sys.argv:
        port = int(sys.argv[sys.argv.index('--port') + 1])
    app.run(debug=True, port=port, threaded=True)
