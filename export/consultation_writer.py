"""Generate a consultation/groundtruth Excel for doctor review.

Layout:
  Col A: Field Key
  Col B: Field Header (human-readable)
  Col C: Group
  Col D: Current Type
  Col E: Current Allowed Values
  Col F: All Unique Values Found
  Col G: LLM Suggested Type
  Col H: LLM Suggested Values
  Col I onwards: One column per patient (header = initials), showing extracted value
  Last 2 cols: Doctor's Type, Doctor's Values (blank for doctor to fill)

Cells with no value show "VALUE NOT FOUND" in red.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from config import get_all_fields, get_groups, load_overrides


_RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
_RED_FONT = Font(color='9C0006', bold=True, size=9)
_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
_DOCTOR_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
_DOCTOR_FONT = Font(bold=True, size=10, color='375623')


def _suggest_type(values: list[str], current_type: str) -> str:
    """Suggest a field type based on observed values."""
    if not values:
        return current_type
    # Check if all values are dates
    import re
    date_pattern = re.compile(r'^\d{1,2}/\d{1,2}/\d{2,4}$')
    if all(date_pattern.match(v) for v in values):
        return 'date'
    # Check if all values are numeric
    try:
        [float(re.sub(r'[^\d.\-]', '', v)) for v in values if v]
        if len(values) >= 1:
            return 'number'
    except (ValueError, TypeError):
        pass
    # Check if few unique values → dropdown
    unique = set(v.strip().lower() for v in values)
    if len(unique) <= 6:
        return 'dropdown'
    # Check for yes/no pattern
    if unique <= {'yes', 'no', 'y', 'n', 'true', 'false'}:
        return 'boolean'
    return 'text'


def _suggest_values(values: list[str]) -> list[str]:
    """Suggest standardised allowed values from observed values."""
    if not values:
        return []
    # Normalise and deduplicate
    normalised = {}
    for v in values:
        key = v.strip().lower()
        if key not in normalised:
            normalised[key] = v.strip()
    # Title-case the most common form
    suggestions = []
    for key, val in normalised.items():
        # Normalise common patterns
        if key in ('positive', '+ve', 'pos', 'yes', 'y'):
            if 'Positive' not in suggestions:
                suggestions.append('Positive')
        elif key in ('negative', '-ve', 'neg', 'no', 'n'):
            if 'Negative' not in suggestions:
                suggestions.append('Negative')
        elif key in ('proficient', 'pmmr', 'mss'):
            if 'Proficient' not in suggestions:
                suggestions.append('Proficient')
        elif key in ('deficient', 'dmmr', 'msi-h', 'msih'):
            if 'Deficient' not in suggestions:
                suggestions.append('Deficient')
        elif key in ('clear', 'r0'):
            if 'Clear' not in suggestions:
                suggestions.append('Clear')
        elif key in ('involved', 'r1'):
            if 'Involved' not in suggestions:
                suggestions.append('Involved')
        elif key in ('threatened',):
            if 'Threatened' not in suggestions:
                suggestions.append('Threatened')
        elif key in ('male', 'm'):
            if 'Male' not in suggestions:
                suggestions.append('Male')
        elif key in ('female', 'f'):
            if 'Female' not in suggestions:
                suggestions.append('Female')
        elif key in ('curative', 'radical'):
            if 'Curative' not in suggestions:
                suggestions.append('Curative')
        elif key in ('palliative',):
            if 'Palliative' not in suggestions:
                suggestions.append('Palliative')
        else:
            suggestions.append(val.title() if len(val) < 30 else val[:30])

    return suggestions[:10]  # cap at 10


def write_consultation_excel(patients: list, output_path: str) -> None:
    """Generate consultation/groundtruth Excel from extraction results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Consultation"

    all_fields = get_all_fields()
    overrides = load_overrides()

    # Collect per-field values across all patients
    field_values: dict[str, list[str]] = {f['key']: [] for f in all_fields}
    for patient in patients:
        for group_name, fields in patient.extractions.items():
            for field_key, fr in fields.items():
                if field_key in field_values and fr.value is not None:
                    field_values[field_key].append(fr.value)

    # Build headers
    fixed_headers = [
        "Field Key", "Field Header", "Group", "Current Type",
        "Current Allowed Values", "All Unique Values",
        "LLM Suggested Type", "LLM Suggested Values",
    ]
    patient_headers = [p.initials or p.id for p in patients]
    doctor_headers = ["Doctor's Type", "Doctor's Values"]

    all_headers = fixed_headers + patient_headers + doctor_headers

    # Write header row
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if header.startswith("Doctor"):
            cell.fill = _DOCTOR_FILL
            cell.font = _DOCTOR_FONT
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Write field rows
    for row_idx, field in enumerate(all_fields, start=2):
        key = field['key']
        override = overrides.get(key, {})
        values = field_values.get(key, [])
        unique_values = sorted(set(values))

        # Fixed columns
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=field['excel_header'])
        ws.cell(row=row_idx, column=3, value=field.get('group_name', ''))
        ws.cell(row=row_idx, column=4, value=override.get('type', field['type']))
        ws.cell(row=row_idx, column=5, value=', '.join(override.get('allowed_values', [])))
        ws.cell(row=row_idx, column=6, value=', '.join(unique_values))
        ws.cell(row=row_idx, column=7, value=_suggest_type(unique_values, field['type']))
        ws.cell(row=row_idx, column=8, value=', '.join(_suggest_values(unique_values)))

        # Per-patient columns
        for p_idx, patient in enumerate(patients):
            col = len(fixed_headers) + p_idx + 1
            # Find this field's value for this patient
            value = None
            for group_name, fields in patient.extractions.items():
                fr = fields.get(key)
                if fr and fr.value is not None:
                    value = fr.value
                    break

            cell = ws.cell(row=row_idx, column=col)
            if value is not None:
                cell.value = value
                cell.font = Font(size=9)
            else:
                cell.value = "VALUE NOT FOUND"
                cell.fill = _RED_FILL
                cell.font = _RED_FONT

        # Doctor columns (blank)
        doc_type_col = len(fixed_headers) + len(patients) + 1
        doc_vals_col = len(fixed_headers) + len(patients) + 2
        ws.cell(row=row_idx, column=doc_type_col).fill = _DOCTOR_FILL
        ws.cell(row=row_idx, column=doc_vals_col).fill = _DOCTOR_FILL

    # Auto-width first 8 columns
    for col_idx in range(1, len(fixed_headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(ws.max_row, 10)):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 35)

    # Patient columns narrower
    for col_idx in range(len(fixed_headers) + 1, len(all_headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15

    # Freeze panes: first row + first 2 columns
    ws.freeze_panes = 'C2'

    wb.save(output_path)


def import_consultation_excel(file_path: str) -> dict:
    """Import doctor-filled consultation Excel back into field overrides.

    Reads the 'Doctor's Type' and 'Doctor's Values' columns and returns
    a dict suitable for save_overrides().
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    ws = wb.active

    # Find column indices by header name
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def col_idx(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    key_col = col_idx("Field Key")
    doc_type_col = col_idx("Doctor's Type")
    doc_vals_col = col_idx("Doctor's Values")

    if key_col is None:
        return {}

    overrides = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        field_key = row[key_col] if key_col < len(row) else None
        if not field_key:
            continue

        doc_type = row[doc_type_col] if doc_type_col is not None and doc_type_col < len(row) else None
        doc_vals = row[doc_vals_col] if doc_vals_col is not None and doc_vals_col < len(row) else None

        if doc_type or doc_vals:
            override = {}
            if doc_type:
                override['type'] = str(doc_type).strip()
            if doc_vals:
                override['allowed_values'] = [v.strip() for v in str(doc_vals).split(',') if v.strip()]
            overrides[field_key] = override

    wb.close()
    return overrides
