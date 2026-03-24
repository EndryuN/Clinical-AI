import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from config import get_all_fields, get_groups


# Confidence basis → (fill hex, font italic, font colour)
_BASIS_STYLES = {
    "structured_verbatim": ("C6EFCE", False, None),      # green
    "freeform_verbatim":   ("FFEB9C", True,  "9C6500"),  # orange + italic
    "freeform_inferred":   ("FFC7CE", False, "9C0006"),  # red
    "edited":              ("D9D9D9", False, None),       # grey
    "absent":              (None,     False, None),       # no fill
}


def _get_group_colors() -> dict:
    colors = {}
    for group in get_groups():
        hex_color = group.get('color', '#D9D9D9').lstrip('#')
        for field in group['fields']:
            colors[field['key']] = hex_color
    return colors


def write_excel(patients: list, output_path: str, source_name: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prototype V1"

    all_fields = get_all_fields()
    group_colors = _get_group_colors()
    OFFSET = 1  # unique_id occupies column 1; all schema fields shift right by 1

    # Column 1: unique_id header
    uid_header = ws.cell(row=1, column=1, value="unique_id")
    uid_header.font = Font(bold=True, size=9)

    # Field headers (shifted by OFFSET)
    header_font = Font(bold=True, size=9)
    for field in all_fields:
        col = field['excel_column'] + OFFSET
        cell = ws.cell(row=1, column=col, value=field['excel_header'])
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        hex_color = group_colors.get(field['key'], 'D9D9D9')
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    # Patient data rows
    for row_idx, patient in enumerate(patients, start=2):
        ws.cell(row=row_idx, column=1, value=patient.unique_id or patient.id)

        for group_name, fields in patient.extractions.items():
            for field_key, fr in fields.items():
                col = _get_column(all_fields, field_key)
                if col is None:
                    continue
                col += OFFSET
                if fr.value is not None:
                    cell = ws.cell(row=row_idx, column=col, value=fr.value)
                    field_def = _get_field_def(all_fields, field_key)
                    if field_def and field_def['type'] == 'date':
                        cell.number_format = 'DD/MM/YYYY'

                    fill_hex, italic, font_color = _BASIS_STYLES.get(
                        fr.confidence_basis, (None, False, None)
                    )
                    if fill_hex:
                        cell.fill = PatternFill(
                            start_color=fill_hex, end_color=fill_hex, fill_type='solid'
                        )
                    if italic or font_color:
                        cell.font = Font(
                            italic=italic,
                            color=font_color or "000000"
                        )
                    # Cell comment for edited originals
                    if fr.edited and fr.original_value is not None:
                        cell.comment = Comment(
                            f"Original: {fr.original_value}", "MDT Extractor"
                        )

    # Legend
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    for i, (basis, label) in enumerate([
        ("structured_verbatim", "Green — extracted from structured field (high confidence)"),
        ("freeform_verbatim",   "Orange — found verbatim in freeform text (medium confidence)"),
        ("freeform_inferred",   "Red — inferred by LLM, not verbatim (low confidence)"),
        ("edited",              "Grey — manually edited by clinician"),
    ], start=1):
        fill_hex, italic, font_color = _BASIS_STYLES[basis]
        c = ws.cell(row=legend_row + i, column=1, value=label)
        if fill_hex:
            c.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
        if italic or font_color:
            c.font = Font(italic=italic, color=font_color or "000000")

    # Auto-width (cap at 30)
    for col_idx in range(1, len(all_fields) + OFFSET + 1):
        max_len = 0
        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(ws.max_row, 5)
        ):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 2, 30)

    # ── Metadata sheet ──
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.sheet_state = 'hidden'
    ws_meta.append(["SOURCE_FILE", source_name])
    META_HEADERS = [
        "unique_id", "field_key", "confidence_basis", "reason",
        "source_cell_row", "source_cell_col", "source_snippet",
        "edited", "original_value", "coverage_pct"
    ]
    ws_meta.append(META_HEADERS)

    for patient in patients:
        pid = patient.unique_id or patient.id
        for group_name, fields in patient.extractions.items():
            for field_key, fr in fields.items():
                sc_row = fr.source_cell['row'] if fr.source_cell else None
                sc_col = fr.source_cell['col'] if fr.source_cell else None
                snippet = (fr.source_snippet or '')[:200]
                ws_meta.append([
                    pid,
                    field_key,
                    fr.confidence_basis,
                    fr.reason or '',
                    sc_row,
                    sc_col,
                    snippet,
                    str(fr.edited).lower(),
                    fr.original_value or '',
                    patient.coverage_pct,
                ])

    # ── RawCells sheet ──
    ws_rc = wb.create_sheet("RawCells")
    ws_rc.sheet_state = 'hidden'
    ws_rc.append(["unique_id", "row", "col", "text", "coverage_json"])

    for patient in patients:
        pid = patient.unique_id or patient.id
        for cell in patient.raw_cells:
            r, c_idx = cell['row'], cell['col']
            cell_key = f"{r},{c_idx}"
            spans = patient.coverage_map.get(cell_key, [])
            ws_rc.append([
                pid,
                r,
                c_idx,
                cell.get('text', ''),
                json.dumps(spans),
            ])

    wb.save(output_path)


def _get_column(all_fields: list, key: str) -> int | None:
    for f in all_fields:
        if f['key'] == key:
            return f['excel_column']
    return None


def _get_field_def(all_fields: list, key: str) -> dict | None:
    for f in all_fields:
        if f['key'] == key:
            return f
    return None
