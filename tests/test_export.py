import os
import tempfile
import pytest
from openpyxl import load_workbook
from models import PatientBlock, FieldResult
from export.excel_writer import write_excel


def test_excel_exports_metadata_sheet():
    patient = PatientBlock(
        id="p001", initials="AO", nhs_number="9990000001", raw_text="test",
        extractions={
            "Demographics": {
                "dob": FieldResult(value="26/05/1970", confidence_basis="structured_verbatim", reason="regex match"),
                "initials": FieldResult(value="AO", confidence_basis="structured_verbatim", reason=""),
                "mrn": FieldResult(value="9990001", confidence_basis="structured_verbatim", reason=""),
                "nhs_number": FieldResult(value="9990000001", confidence_basis="freeform_verbatim", reason="LLM inferred"),
                "gender": FieldResult(value="Male", confidence_basis="structured_verbatim", reason=""),
                "previous_cancer": FieldResult(value=None, confidence_basis="absent", reason=""),
                "previous_cancer_site": FieldResult(value=None, confidence_basis="absent", reason=""),
            }
        }
    )
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        write_excel([patient], path)
        wb = load_workbook(path)
        assert "Metadata" in wb.sheetnames
        ws = wb["Metadata"]
        # Row 1 is SOURCE_FILE global metadata
        assert ws.cell(1, 1).value == "SOURCE_FILE"
        # Row 2 has named headers
        headers = [ws.cell(2, c).value for c in range(1, 11)]
        assert "unique_id" in headers
        assert "confidence_basis" in headers
        assert "field_key" in headers
        # Find nhs_number row (data starts at row 3)
        cb_idx = headers.index("confidence_basis")
        fk_idx = headers.index("field_key")
        reason_idx = headers.index("reason")
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        nhs_row = next((r for r in rows if r[fk_idx] == "nhs_number"), None)
        assert nhs_row is not None
        assert nhs_row[cb_idx] == "freeform_verbatim"
        assert "LLM inferred" in (nhs_row[reason_idx] or "")
        wb.close()
    finally:
        os.unlink(path)


@pytest.mark.skip(reason="Requires _import_excel update (Task 8) for new column layout")
def test_excel_round_trip_restores_confidence_and_reason():
    """Export then re-import — confidence and reason must survive."""
    import sys
    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
    from app import _import_excel  # noqa: import after path setup
    patient = PatientBlock(
        id="p001", initials="AO", nhs_number="9990000001", raw_text="test",
        extractions={
            "Demographics": {
                "dob": FieldResult(value="26/05/1970", confidence_basis="structured_verbatim", reason="regex"),
                "initials": FieldResult(value="AO", confidence_basis="structured_verbatim", reason=""),
                "mrn": FieldResult(value="9990001", confidence_basis="structured_verbatim", reason=""),
                "nhs_number": FieldResult(value="9990000001", confidence_basis="freeform_verbatim", reason="LLM"),
                "gender": FieldResult(value="Male", confidence_basis="structured_verbatim", reason=""),
                "previous_cancer": FieldResult(value=None, confidence_basis="absent", reason=""),
                "previous_cancer_site": FieldResult(value=None, confidence_basis="absent", reason=""),
            }
        }
    )
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        write_excel([patient], path)
        reimported = _import_excel(path)
        demo = reimported[0].extractions.get("Demographics", {})
        assert demo["nhs_number"].confidence == "medium"
        assert "LLM" in (demo["nhs_number"].reason or "")
        assert demo["dob"].confidence == "high"
    finally:
        os.unlink(path)


def test_excel_round_trip():
    patient = PatientBlock(
        id="patient_001",
        initials="AO",
        nhs_number="9990000001",
        raw_text="test",
        extractions={
            "Demographics": {
                "dob": FieldResult(value="26/05/1970", confidence_basis="structured_verbatim"),
                "initials": FieldResult(value="AO", confidence_basis="structured_verbatim"),
                "mrn": FieldResult(value="9990001", confidence_basis="structured_verbatim"),
                "nhs_number": FieldResult(value="9990000001", confidence_basis="structured_verbatim"),
                "gender": FieldResult(value="Male", confidence_basis="structured_verbatim"),
                "previous_cancer": FieldResult(value="No", confidence_basis="freeform_verbatim"),
                "previous_cancer_site": FieldResult(value="N/A", confidence_basis="freeform_inferred"),
            }
        }
    )

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name

    try:
        write_excel([patient], output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.title == "Prototype V1"
        # Column 1 is unique_id, schema fields shifted by OFFSET=1
        assert ws.cell(row=1, column=1).value == "unique_id"
        assert ws.cell(row=2, column=3).value == "AO"      # initials: excel_column 2 + OFFSET 1 = 3
        assert ws.cell(row=2, column=6).value == "Male"     # gender: excel_column 5 + OFFSET 1 = 6
        wb.close()
    finally:
        os.unlink(output_path)


def _make_patient_with_rawcells():
    return PatientBlock(
        id="9990001",
        unique_id="07032025_AO_M_9990001",
        initials="AO",
        nhs_number="9990000001",
        gender="Male",
        mdt_date="07/03/2025",
        raw_cells=[
            {"row": 0, "col": 0, "text": "Patient Details"},
            {"row": 1, "col": 0, "text": "Hospital Number: 9990001"},
            {"row": 5, "col": 0, "text": "T3 tumour at 5cm. EMVI positive."},
        ],
        coverage_map={
            "5,0": [{"start": 0, "end": 15, "used": True},
                    {"start": 15, "end": 33, "used": False}]
        },
        coverage_pct=45.5,
        extractions={
            "Demographics": {
                "mrn": FieldResult(value="9990001", confidence_basis="structured_verbatim"),
                "initials": FieldResult(value="AO", confidence_basis="structured_verbatim"),
                "dob": FieldResult(value="01/01/1970", confidence_basis="structured_verbatim"),
                "nhs_number": FieldResult(value="9990000001", confidence_basis="freeform_verbatim"),
                "gender": FieldResult(value="Male", confidence_basis="structured_verbatim"),
                "previous_cancer": FieldResult(value=None, confidence_basis="absent"),
                "previous_cancer_site": FieldResult(value=None, confidence_basis="absent"),
            }
        }
    )


def test_excel_has_rawcells_sheet():
    patient = _make_patient_with_rawcells()
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        write_excel([patient], path)
        wb = load_workbook(path)
        assert "RawCells" in wb.sheetnames
        ws_rc = wb["RawCells"]
        headers = [ws_rc.cell(1, c).value for c in range(1, 6)]
        assert headers == ["unique_id", "row", "col", "text", "coverage_json"]
        wb.close()
    finally:
        os.unlink(path)


def test_excel_rawcells_contains_patient_data():
    patient = _make_patient_with_rawcells()
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        write_excel([patient], path)
        wb = load_workbook(path)
        ws_rc = wb["RawCells"]
        rows = list(ws_rc.iter_rows(min_row=2, values_only=True))
        cell_row = next((r for r in rows if r[1] == 5 and r[2] == 0), None)
        assert cell_row is not None
        assert "T3 tumour" in (cell_row[3] or "")
        import json
        spans = json.loads(cell_row[4])
        assert isinstance(spans, list)
        wb.close()
    finally:
        os.unlink(path)


def test_excel_unique_id_is_first_column():
    patient = _make_patient_with_rawcells()
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        write_excel([patient], path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "unique_id"
        assert ws.cell(2, 1).value == "07032025_AO_M_9990001"
        wb.close()
    finally:
        os.unlink(path)


def test_excel_metadata_sheet_has_confidence_basis_column():
    patient = _make_patient_with_rawcells()
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        write_excel([patient], path)
        wb = load_workbook(path)
        ws_meta = wb["Metadata"]
        headers = [ws_meta.cell(2, c).value for c in range(1, 11)]
        assert "confidence_basis" in headers
        assert "unique_id" in headers
        rows = list(ws_meta.iter_rows(min_row=3, values_only=True))
        col_idx = headers.index("confidence_basis")
        key_idx = headers.index("field_key")
        nhs_row = next((r for r in rows if r[key_idx] == "nhs_number"), None)
        assert nhs_row is not None
        assert nhs_row[col_idx] == "freeform_verbatim"
        wb.close()
    finally:
        os.unlink(path)
